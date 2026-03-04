#!/usr/bin/env python3
"""
tag_library.py — Rekordbox tagging system for Lionel's DJ library
Version 6.0 | March 2026

Scans all audio files in DJ_MUSIC and produces:
  1. rekordbox_import.xml  — Import into Rekordbox:
                              File → Library → Import Library → Select File
  2. library_tags.xlsx     — Master spreadsheet of all tracks + tags
  3. [optional] Writes Genre, Comment (sub-genre | energy), Year to ID3 tags

Usage:
  python3 tag_library.py               # Generate XML + xlsx only (safe, no file edits)
  python3 tag_library.py --write-tags  # Also write ID3 tags to audio files
  python3 tag_library.py --preview     # Print first 30 rows to terminal, no files written

Energy / Stars system (1–5):
  5★ = Peak energy / floor-filler / anthem
  4★ = High energy / dance floor builder
  3★ = Medium / versatile
  2★ = Lower energy / opener or closer
  1★ = Ambient / tool / percussion loop

Energy Tiers (derived from stars after ENERGY_DOWN / ENERGY_UP modifiers):
  PEAK  = 5★   Floor-filler, anthem, main-stage
  HIGH  = 4★   Dance-floor builder, peak-hour
  MID   = 3★   Versatile, set-filler
  LOW   = 1–2★ Opener, closer, ambient

Rekordbox Color Labels:
  Red    = Peak energy / crowd-pleaser
  Orange = High energy / dance floor builder
  Yellow = Mid-energy / Israeli / Classics / versatile
  Pink   = Fun / pop / feel-good / cover
  Blue   = Emotional / rock / introspective
  Purple = Dark / late-night / techno / psych
  Aqua   = World / ambient / ethereal
  Green  = Tools / loops / utility / reggae

v6.0 changes (Electronic sub-genre deep-dive — target <30 unclassified):
  - Electronic: +3 new rules: Tech House, Melodic House, Club/Dance Pop
  - Electronic: 10 expanded rules targeting 175 of 187 previously unclassified
    tracks (Uplifting Trance +10 artists, Nu-Disco +2, Techno +1, Global Bass +2,
    UK Dance +1 variant, Breakbeat +5, Bass/Dubstep +10, French House +4,
    Downtempo +8, Electro +8, Lo-Fi +8, Indie Electronic +21)

v5.0 changes (Energy scoring + Electronic taxonomy expansion):
  - Energy tier scoring: classify_track() now returns (subgenre, stars, color, energy)
    STAR_TO_ENERGY: 5→"PEAK", 4→"HIGH", 3→"MID", 2→"LOW", 1→"LOW"
    Derived AFTER ENERGY_DOWN / ENERGY_UP star modifiers are applied.
    Stored as "SubGenre | Energy" in rekordbox Comments; new Energy column in xlsx.
  - Electronic: +12 new sub-genre rules targeting ~60–70% of 279 unclassified tracks:
      Progressive Trance   (Chicane, Omnia, Rodg, Estiva, Kryder, Genix, Delerium)
      Goa / Ethnic Psy     (1200 Micrograms, Etnica, Filteria, Kino Todo, Outsiders)
      Major Lazer / Global Bass (Major Lazer, Diplo, Jack Ü, TroyBoi)
      Commercial Dance     (Robin Schulz, Tungevaag, Tim Berg, Alesso)
      Progressive House    (Deadmau5, Noizu, Gorje Hewek, Max Cooper)
      Nu-Disco / Funky House (Mousse T, Martin Solveig, Klingande, Polo & Pan, C2C)
      UK Dance             (Jax Jones, Sigala, MK, Route 94, Sam Feldt)
      New Wave Electronic  (New Order, Joy Division, Depeche Mode — Electronic folder)
      Synth-Pop / Electropop (La Roux, M83, Chvrches, Stromae, Cut Copy)
      Electro / Krautrock  (Kraftwerk, Gesaffelstein, Moderat, SBTRKT)
      Trip-Hop             (Moby, Portishead, Massive Attack, Tricky, Unkle, Zero 7)
      Organic Bass         (Polish Ambassador, Equanimous, Mfinity, Kalya Scintilla)
  - Electronic: 4 expanded rules:
      Breakbeat  — added "the_chemical_brothers" (underscore filename fix)
      Future Bass — added Flume, Illenium, Big Wild, Medasin
      Downtempo/Chillout — added Khruangbin, Kerala Dust, Nils Frahm
      Indie Electronic — added The XX, Beach House, Cigarettes After Sex, Alvvays

v4.0 changes (Architectural taxonomy upgrade from Google Doc framework):
  - Israeli:    + Mediterranean Trap (Mergui, Ron Nesher, Eden Hason, Noga Erez,
                  Anna Zak) — 808-heavy + Hebrew/Mediterranean melisma
                + Ethno-Fusion (A-WA, Yemen Blues, Idan Raichel Project)
                  — global bass + traditional roots; now separate from Folk
                + Eurovision Pop (Netta, Nadav Guedj, Noa Kirel Eurovision)
                  — international-stage commercial dance-pop
                + Nitzhonot (uniquely Israeli Goa trance, 145–155 BPM,
                  oriental melodies / "laserkicks")
                + Expanded Israeli Hip-Hop: Shabak Samech, Fishi Ha-Gadol,
                  Hadag Nahash, Jimbo J, Ravid Plotnik, Tuna, Mook E
                + Renamed "Israeli Folk" → "Shirei Eretz Yisrael"
                  (Chava Alberstein, Naomi Shemer, early Shlomo Artzi)
  - Electronic: + Psytrance split into: Full-On Psy (Infected Mushroom, Astrix,
                  Captain Hook, Vibe Tribe, 140–148 BPM), Dark/Forest Psy
                  (organic/gritty, 150+ BPM), Progressive Psy (128–134 BPM),
                  Nitzhonot (Israeli Goa oriental, 145–155 BPM),
                  Zenonesque/Psy-Tech (minimal techno-psy crossover)
                + Amapiano (log-drum basslines, 110–115 BPM, South African)
                + Afro Tech (Black Coffee, Shimza — tribal + futuristic synths)
                + Hyperpop / Krushclub (glitchy, high-velocity pop/industrial)
  - House:      + Afro House / Afro Tech artists expanded (Black Coffee, Shimza)
  - World:      + Amapiano in World as well for South African dance floor context
"""

import argparse
import re
import sys
import unicodedata
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import date
from pathlib import Path

# ──────────────────────────────────────────────────────────────────────────────
# PATHS
# ──────────────────────────────────────────────────────────────────────────────
VM_DJMUSIC  = Path("/sessions/busy-funny-noether/mnt/Music/DJ_MUSIC")
MAC_DJMUSIC = "/Users/Lionmit/Music/DJ_MUSIC"
OUT_DIR     = Path("/sessions/busy-funny-noether/mnt/Music")
XML_OUT     = OUT_DIR / "rekordbox_import.xml"
XLSX_OUT    = OUT_DIR / "library_tags.xlsx"
AUDIO_EXTS  = {".mp3", ".wav", ".flac", ".m4a", ".aiff", ".aif", ".ogg"}
TODAY       = date.today().isoformat()

# ──────────────────────────────────────────────────────────────────────────────
# REKORDBOX CONSTANTS
# ──────────────────────────────────────────────────────────────────────────────
COLOR_NONE   = 0
COLOR_PINK   = 1
COLOR_RED    = 2
COLOR_ORANGE = 3
COLOR_YELLOW = 4
COLOR_GREEN  = 5
COLOR_AQUA   = 6
COLOR_BLUE   = 7
COLOR_PURPLE = 8

COLOR_NAMES = {
    0: "—", 1: "Pink", 2: "Red", 3: "Orange",
    4: "Yellow", 5: "Green", 6: "Aqua", 7: "Blue", 8: "Purple"
}

STARS_TO_RATING = {0: 0, 1: 51, 2: 102, 3: 153, 4: 204, 5: 255}

EXT_TO_KIND = {
    ".mp3": "MP3 File", ".wav": "WAV File",  ".flac": "FLAC File",
    ".m4a": "M4A File", ".aiff": "AIFF File", ".aif": "AIFF File",
    ".ogg": "OGG File",
}

YEAR_RE = re.compile(r'\b(19[5-9]\d|20[0-2]\d)\b')

# ──────────────────────────────────────────────────────────────────────────────
# GENRE CONFIGURATION
# Keyed by the START of the folder name (case-insensitive prefix match)
# Value: (genre_label, default_stars, default_color, default_subgenre)
# ──────────────────────────────────────────────────────────────────────────────
GENRE_CONFIG = [
    ("00_inbox",    "INBOX",        3, COLOR_NONE,   "Unclassified"),
    ("01 israeli",  "Israeli",      3, COLOR_YELLOW, "Israeli Pop"),
    ("02 hip-hop",  "Hip-Hop & R&B",4, COLOR_ORANGE, "Hip-Hop"),
    ("03 house",    "House & Dance",4, COLOR_RED,    "House"),
    ("04 electron", "Electronic",   3, COLOR_PURPLE, "Electronic"),
    ("05 pop",      "Pop",          3, COLOR_PINK,   "Pop"),
    ("06 rock",     "Rock",         3, COLOR_BLUE,   "Rock"),
    ("07 latin",    "Latin",        4, COLOR_ORANGE, "Latin"),
    ("08 classic",  "Classics",     3, COLOR_YELLOW, "Classic"),
    ("09 world",    "World",        2, COLOR_AQUA,   "World"),
    ("10 tools",    "Tools & FX",   1, COLOR_GREEN,  "Tool"),
    ("11 remixes",  "Remixes",      3, COLOR_PINK,   "Remix"),
]

# ──────────────────────────────────────────────────────────────────────────────
# SUB-GENRE RULES  (v4.0 — Architectural taxonomy upgrade)
# Format: ([keyword_list], sub_genre_name, stars, color)
# Matched against lowercased filename. First match wins; None = use default.
# ──────────────────────────────────────────────────────────────────────────────
SUBGENRE_RULES = {

    # ── ISRAELI ──────────────────────────────────────────────────────────────
    # Rule order is critical (first-match-wins):
    #   1. Nitzhonot first (before Israeli Electronic, to claim Israeli Goa artists)
    #   2. Mediterranean Trap (before Israeli Pop, to claim Mergui / Noga Erez)
    #   3. Ethno-Fusion (before Shirei Eretz Yisrael, to claim Idan Raichel)
    #   4. Eurovision Pop (before Israeli Pop, to claim Netta / Nadav Guedj)
    #   5. Israeli Hip-Hop (before Mizrahi, to claim Trap keyword)
    #   6. Mizrahi, Israeli Pop, Israeli Electronic, Israeli Rock
    #   7. Shirei Eretz Yisrael (catch-all Folk / Classic Israeli)
    "Israeli": [
        # Nitzhonot — uniquely Israeli Goa trance; oriental melodies + "laserkicks"
        # 145–155 BPM. Must precede Israeli Electronic so Goa artists don't fall through.
        (["nitzhonot", "nitzarim", "nitzo",
          "infected mushroom - ", "atomic pulse",
          "duvdev", "yahel", "skazi",
          "psytrance israel", "goa israel",
          "psy israel"],                               "Nitzhonot",           5, COLOR_RED),
        # Mediterranean Trap — 808-heavy + Hebrew/Mediterranean melisma
        # (Google Doc: Mergui, Noga Erez, Anna Zak, Ron Nesher, Eden Hason)
        (["mergui", "ron nesher", "eden hason",
          "anna zak", "noga erez",
          "med trap", "mediterranean trap"],           "Mediterranean Trap",  4, COLOR_ORANGE),
        # Ethno-Fusion — global bass + traditional Middle Eastern/Yemenite roots
        # (Google Doc: A-WA, Yemen Blues, Idan Raichel Project)
        (["a-wa", "a wa ", "arava - ", "habib galbi",
          "yemen blues", "idan raichel",
          "ethnix ", "ethno-fusion", "ethno fusion"],  "Ethno-Fusion",        3, COLOR_AQUA),
        # Eurovision Pop — international-stage commercial Israeli dance-pop
        # (Google Doc: Netta, Nadav Guedj, Eurovision context Noa Kirel)
        (["netta", "nadav guedj", "noa kirel - ",
          "noa kirel feat", "noa kirel x",
          "zingzillas", "toy - ", "chocolate",
          "eurovision", "israel - eurovision"],        "Eurovision Pop",      4, COLOR_PINK),
        # Israeli Hip-Hop — expanded (Google Doc adds Shabak Samech, Fishi Ha-Gadol,
        # Hadag Nahash, Jimbo J, Ravid Plotnik, Tuna + existing Mook E, Subliminal)
        (["rap", "hip hop", "hip-hop", "טראפ",
          "subliminal", "hadag nachash", "hadag nahash",
          "mook e", "shabak samech", "שבק סמך",
          "fishi ha-gadol", "fishi hagadol", "פישי הגדול",
          "jimbo j", "ג'ימבו",
          "ravid plotnik", "רביד פלוטניק",
          "tuna - ", "טונה",
          "echo - ", "riff cohen",
          "israeli rap", "israeli hip-hop"],           "Israeli Hip-Hop",     4, COLOR_ORANGE),
        # Mizrahi — Andalusian/Yemenite/Oriental tradition; most recognisable sub-genre
        (["mizrahi", "מזרחי", "oriental", "פיוט",
          "eyal golan", "sarit hadad", "moshe peretz", "kobi peretz",
          "eden ben zaken", "zehava ben", "ofra haza",
          "boaz sharabi", "zohar argov"],              "Mizrahi",             3, COLOR_YELLOW),
        # Israeli Pop — modern pop era (Omer Adam generation)
        # Noa Kirel catch-all here (Eurovision-specific already caught above)
        (["omer adam", "noa kirel",
          "shiri maimon",
          "marina maximilian", "static & ben el", "static ben el",
          "dana international", "ivri lider", "harel skaat",
          "rita - "],                                  "Israeli Pop",         3, COLOR_YELLOW),
        # Israeli Electronic — Tel Aviv club scene (general EDM / non-Goa)
        (["electronic", "edm", "made in tlv", "galala",
          "captain hook", "astrix", "vibe tribe"],     "Israeli Electronic",  4, COLOR_ORANGE),
        # Israeli Rock
        (["rock", "metal", "alternative",
          "shlomo artzi", "yehuda poliker", "ethnix",
          "mashina", "kaveret", "rockfour"],           "Israeli Rock",        3, COLOR_BLUE),
        # Shirei Eretz Yisrael — classic Israeli folk canon
        # (Google Doc: rename from "Israeli Folk"; Chava Alberstein, Naomi Shemer)
        (["folk", "עממי", "traditional",
          "naomi shemer", "chava alberstein", "חוה אלברשטיין",
          "corinne allal", "להקת פיקוד", "shirei eretz",
          "eretz yisrael"],                            "Shirei Eretz Yisrael",2, COLOR_AQUA),
    ],

    # ── HIP-HOP & R&B ────────────────────────────────────────────────────────
    "Hip-Hop & R&B": [
        # Phonk — Memphis-derived dark trap; TikTok drift culture
        (["phonk", "drift phonk", "memphis rap"],       "Phonk",              4, COLOR_PURPLE),
        # Drill — high-energy dark rap; UK Drill
        (["pop smoke", "sheff g", "fivio",
          "uk drill", "drill"],                         "Drill",              5, COLOR_RED),
        # Trap — hard trap (drill keyword moved to Drill rule above)
        (["trap", "future - "],                         "Trap",               5, COLOR_RED),
        # R&B / Soul — expanded with new-school R&B artists
        (["r&b", "soul",
          "alicia keys", "usher", "beyoncé", "beyonce",
          "rihanna", "mariah", "ring the alarm", "scream",
          "empire state", "kiss me more", "doja cat",
          "khalid", "sza", "h.e.r.", "summer walker",
          "the weeknd", "frank ocean", "miguel",
          "bryson tiller"],                             "R&B",                3, COLOR_PINK),
        # Boom Bap / Golden Age
        (["a tribe called quest", "tribe called quest",
          "beastie", "biggie", "2pac", "tupac", "dr. dre",
          "n.w.a", "nwa", "public enemy", "heavy d", "ll cool j",
          "ice cube", "ice-t", "cam'ron", "rza", "wu-tang",
          "nas - ", "jay-z", "jay z", "notorious b.i.g",
          "pete rock", "de la soul", "gang starr"],     "Boom Bap",           4, COLOR_ORANGE),
        # G-Funk — West Coast Parliament-inspired groove
        (["snoop dogg", "warren g", "nate dogg",
          "kurupt", "daz dillinger", "g-funk", "g funk"],
                                                        "G-Funk",             4, COLOR_ORANGE),
        # Old School Hip-Hop — pre-golden age; breakdance era
        (["run-dmc", "run dmc", "grandmaster flash",
          "afrika bambaataa", "kurtis blow",
          "sugarhill gang", "sugar hill gang",
          "old school hip-hop"],                        "Old School Hip-Hop", 3, COLOR_YELLOW),
        # Lo-Fi Hip-Hop — study beats; chill sets
        (["j dilla", "nujabes", "madlib",
          "lofi hip", "lo-fi hip", "lo-fi beats"],      "Lo-Fi Hip-Hop",      2, COLOR_AQUA),
        # New School / Contemporary
        (["drake", "kendrick", "j. cole", "j cole", "chance",
          "travis scott", "cardi b", "nicki minaj", "post malone",
          "21 savage", "lil uzi", "roddy ricch"],       "New School Hip-Hop", 4, COLOR_ORANGE),
        (["moombah", "dancehall"],                      "Hip-Hop/Dance",      4, COLOR_ORANGE),
        (["remix", "mashup", " vs ", "bootleg"],        "Hip-Hop Remix",      3, COLOR_PINK),
    ],

    # ── HOUSE & DANCE ────────────────────────────────────────────────────────
    "House & Dance": [
        # Big Room / Festival — highest energy; main stage anthems
        (["avicii", "hardwell", "afrojack", "martin garrix",
          "dimitri vegas", "like mike", "w&w", "axwell",
          "big room", "festival mix", "festival edit",
          "mainstage", "ultra mix"],                    "Big Room/Festival",  5, COLOR_RED),
        # Uplifting Trance — confirmed in library
        (["armin van buuren", "above & beyond", "above and beyond",
          "gareth emery", "cosmic gate", "paul van dyk",
          "tiesto", "ferry corsten", "markus schulz",
          "uplifting", "trance anthem"],                "Uplifting Trance",   5, COLOR_ORANGE),
        # Hardstyle — Dutch hard dance; festival peak
        (["headhunterz", "brennan heart", "wildstylez",
          "noisecontrollers", "zatox", "hardstyle"],    "Hardstyle",          5, COLOR_RED),
        # Melodic House & Techno — lane 8 / Tale Of Us vibe
        (["lane 8", "nils hoffmann", "adriatique",
          "tale of us", "innellea",
          "rufus du sol", "rüfüs du sol",
          "melodic house", "melodic techno"],           "Melodic House",      4, COLOR_BLUE),
        # Future House — Tchami era future bass house
        (["tchami", "oliver $", "don diablo",
          "future house"],                              "Future House",       4, COLOR_ORANGE),
        # Tech House — Fisher / Chris Lake era
        (["tech house", "tech-house",
          "fisher", "chris lake", "will clarck",
          "anna - ", "bart skils", "joseph capriati"],  "Tech House",         4, COLOR_ORANGE),
        # Funky / Jackin' House
        (["funky house", "jackin house", "jackin",
          "funky "],                                    "Funky/Jackin' House",4, COLOR_ORANGE),
        # Deep House — Moodymann / Larry Heard roots + keywords
        (["deep house", "deep ",
          "moodymann", "kerri chandler", "larry heard"], "Deep House",        3, COLOR_BLUE),
        # UK Garage — 2-step; Craig David era
        (["craig david", "artful dodger", "mj cole",
          "uk garage", "2-step", "2step", "so solid"],  "UK Garage",          3, COLOR_AQUA),
        # Piano House
        (["piano house", "piano "],                     "Piano House",        3, COLOR_YELLOW),
        # Vocal House
        (["gorgon city", "camelphat", "lost frequencies",
          "vocal house", "vocal", "kia love", "vula"],  "Vocal House",        3, COLOR_PINK),
        # Progressive House
        (["progressive house", "progressive"],          "Progressive House",  4, COLOR_RED),
        # Eurodance — 90s dance pop
        (["eurodance", "euro ", "snap!", "dj bobo", "haddaway",
          "2 unlimited", "gala", "dr. alban", "ace of base",
          "corona", "la bouche"],                       "Eurodance",          4, COLOR_ORANGE),
        # Afro House / Afro Tech — Black Coffee, Shimza (Google Doc upgrade)
        (["afro house", "afro tech",
          "black coffee", "shimza",
          "masters at work", "afrobeat house"],         "Afro House",         4, COLOR_ORANGE),
        # Nu-Disco / Disco House
        (["nu-disco", "nu disco", "disco"],             "Nu-Disco",           3, COLOR_YELLOW),
        # Trance (generic — catch-all after specific trance rules above)
        (["trance"],                                    "Trance",             4, COLOR_PURPLE),
        # Minimal House
        (["minimal"],                                   "Minimal House",      3, COLOR_PURPLE),
    ],

    # ── ELECTRONIC ──────────────────────────────────────────────────────────
    # v5.0: +12 new sub-genre rules; 4 expanded rules. Rules ordered most-specific
    # first (first-match-wins). Trance → Psy → African/Global → Club/Dance →
    # Synth/New Wave → Swing/Industrial → Hard Floor → Trip-Hop/Chill →
    # House-Adjacent → Indie/Alt.
    "Electronic": [
        # ── TRANCE CLUSTER ─────────────────────────────────────────────────
        # Uplifting Trance — leading rule (large block in library)
        (["armin van buuren", "above & beyond", "above and beyond",
          "andrew rayel", "cosmic gate", "ilan bluestone",
          "denis kenzo", "super8 & tab", "super8 tab",
          "orjan nilsen", "rank1", "david gravell",
          "paul van dyk", "ferry corsten", "markus schulz",
          "gareth emery", "uplifting trance", "a state of trance",
          "asot",
          # v6.0 additions
          "andrew bayer", "alexander popov", "drym", "luke bond",
          "willem de roo", "christina novelli", "arty &", "arty feat",
          "frainbreeze", "feel & ",
          "ashley wallbridge", "simon lee & alvin",
          "omnia feat", "dash berlin"],              "Uplifting Trance",   5, COLOR_ORANGE),
        # Progressive Trance — melodic, vocal-led, 130–138 BPM  [v5.0 NEW]
        (["chicane", "omnia - ", "rodg - ", "estiva",
          "kryder", "genix", "delerium", "conjure one",
          "sunlounger", "shogun - ", "roger shah",
          "progressive trance", "vocal trance",
          "the trainman"],                          "Progressive Trance", 4, COLOR_ORANGE),
        # ── PSYTRANCE CLUSTER ──────────────────────────────────────────────
        # Full-On Psytrance — peak-energy driving psy (140–148 BPM)
        (["infected mushroom", "astrix", "captain hook",
          "vibe tribe", "electric universe",
          "dick trevor", "protoculture",
          "full on psy", "full-on psy",
          "full on psytrance"],                         "Full-On Psy",        5, COLOR_RED),
        # Goa / Ethnic Psy — classic Goa trance  [v5.0 NEW]
        (["1200 micrograms", "etnica", "kino todo",
          "outsiders", "filteria", "hypnoxock",
          "cosmosis", "pleiadians",
          "goa trance", "goa psy", "ethnic psy"],       "Goa/Ethnic Psy",     5, COLOR_RED),
        # Dark / Forest Psytrance — organic, gritty, 150+ BPM
        (["dark psy", "forest psy", "darkpsy",
          "forest psytrance", "dark forest",
          "bardo", "ocelot", "mindwave"],               "Dark/Forest Psy",    5, COLOR_PURPLE),
        # Progressive Psytrance — atmospheric groove, 128–134 BPM
        (["progressive psy", "prog psy", "progpsy",
          "gaudi", "younger brother",
          "shpongle", "entheogenic",
          "progressive goa"],                           "Progressive Psy",    4, COLOR_PURPLE),
        # Nitzhonot — uniquely Israeli Goa trance; oriental melodies + "laserkicks"
        (["nitzhonot", "nitzarim", "nitzachon",
          "goa israel", "psy israel",
          "atomic pulse", "duvdev",
          "nitzaho"],                                   "Nitzhonot",          5, COLOR_RED),
        # Zenonesque / Psy-Tech — minimal, techno-influenced psytrance crossover
        (["zenonesque", "psy-tech", "psytech",
          "zenon records", "zenon ",
          "ott - ", "a forest of stars"],               "Zenonesque/Psy-Tech",4, COLOR_PURPLE),
        # ── AFRICAN / GLOBAL DANCE ─────────────────────────────────────────
        # Amapiano — log-drum basslines, 110–115 BPM, South African
        (["amapiano", "log drum", "log-drum",
          "kabza de small", "dj maphorisa", "focalistic",
          "ami faku", "sun-el musician"],               "Amapiano",           4, COLOR_ORANGE),
        # Afro Tech / Afro House (Electronic flavour)
        (["afro tech", "afro house",
          "black coffee", "shimza",
          "enoo napa", "themba - ",
          "tribal tech"],                               "Afro Tech",          4, COLOR_ORANGE),
        # Major Lazer / Global Bass  [v5.0 NEW]
        (["major lazer", "major_lazer",
          "diplo", "jack \u00fc", "jack u ",
          "walshy fire", "ape drums", "troyboi",
          "troyboy", "biome - ",
          # v6.0
          "naoba", "ahadadream"],                       "Global Bass",        4, COLOR_ORANGE),
        # ── CLUB / DANCE ────────────────────────────────────────────────────
        # Commercial Dance / EDM Pop  [v5.0 NEW]
        (["robin schulz", "robin_schulz", "tungevaag", "tim berg",
          "nicky romero", "otto knows",
          "sebastian ingrosso", "alesso",
          "commercial dance", "edm pop",
          # v6.0
          "mako ", "young franco", "life of dillon",
          "koni &", "sp3ctrum", "tkdjs", "nowifi",
          "snake city", "hxrt", "toby montana",
          "dubdogz"],                                    "Commercial Dance",   4, COLOR_ORANGE),
        # Progressive House  [v5.0 NEW]
        (["deadmau5", "noizu", "gorje hewek",
          "fur coat", "max cooper", "locked groove",
          "hernan cattaneo", "nick warren",
          "progressive house"],                         "Progressive House",  4, COLOR_PURPLE),
        # Nu-Disco / Funky House  [v5.0 NEW]
        (["mousse t", "martin solveig", "klingande",
          "polo & pan", "polo and pan", "c2c - ", "c2c feat",
          "chromeo", "poolside", "crackazat",
          "nu-disco", "nu disco",
          "funky house", "disco house",
          # v6.0
          "stussko"],                                   "Nu-Disco",           4, COLOR_ORANGE),
        # UK Dance  [v5.0 NEW]
        (["jax jones", "jax_jones", "sigala", "mk - ", "mk feat",
          "route 94", "sam feldt", "james hype",
          "secondcity", "tcts",
          "uk dance"],                                  "UK Dance",           4, COLOR_ORANGE),
        # Hyperpop / Krushclub — glitchy, high-velocity pop/industrial fusion
        (["hyperpop", "krushclub", "100 gecs",
          "charli xcx", "bbymutha", "sophie - ",
          "a.g. cook", "oklou",
          "hyper pop", "hypercore"],                    "Hyperpop",           4, COLOR_RED),
        # ── SYNTH / NEW WAVE CLUSTER ─────────────────────────────────────────
        # Synthwave / Retrowave — 80s-inspired cinematic synth
        (["kavinsky", "perturbator", "gunship",
          "carpenter brut", "ollie wride",
          "synthwave", "retrowave", "outrun",
          "darksynth", "newretrowave"],                 "Synthwave/Retrowave",4, COLOR_PURPLE),
        # New Wave / Post-Punk Electronic  [v5.0 NEW]
        # New Order/Joy Division in Electronic folder won't hit Rock's New Wave rule
        (["new order", "joy division",
          "depeche mode", "erasure",
          "pet shop boys", "tears for fears",
          "yazoo", "soft cell",
          "new wave electronic"],                       "New Wave/Post-Punk", 3, COLOR_BLUE),
        # Synth-Pop / Electropop  [v5.0 NEW]
        # La Roux (~5), Stromae (~3) confirmed unclassified in Electronic folder
        (["la roux", "m83", "chvrches", "cut copy",
          "chromatics", "neon indian", "toro y moi",
          "stromae", "magic man",
          "synth pop", "electropop", "electro pop"],    "Synth-Pop",          3, COLOR_BLUE),
        # Electro / Krautrock  [v5.0 NEW]
        # Kraftwerk (~3), Gesaffelstein confirmed unclassified
        (["kraftwerk", "gesaffelstein", "moderat",
          "sbtrkt", "roman flugel", "roman fl\u00fcgel",
          "electroclash", "krautrock",
          # v6.0
          "black strobe", "esg dance",
          "touch and go", "light asylum",
          "celldweller", "neil cicierega",
          "neuroplasm", "carbon decay"],               "Electro",            4, COLOR_PURPLE),
        # ── SWING / INDUSTRIAL ───────────────────────────────────────────────
        # Electro Swing — confirmed: Sim Gretina, Club Des Belugas
        (["electro swing", "electroswing", "sim gretina",
          "caravan palace", "caro emerald", "parov stelar",
          "wanna be like you", "i wanna be like"],      "Electro Swing",      4, COLOR_YELLOW),
        # Industrial / EBM — before Techno to claim "industrial" keyword
        (["ministry - ", "kmfdm", "covenant - ",
          "vnv nation", "front 242",
          "ebm", "industrial metal", "industrial"],     "Industrial/EBM",     4, COLOR_PURPLE),
        # ── HARD FLOOR ───────────────────────────────────────────────────────
        # Techno — dark floor techno
        (["techno", "perc ",
          "adam beyer", "charlotte de witte", "amelie lens",
          "joseph capriati", "spfdj", "nina kraviz",
          "ben klock",
          # v6.0
          "nokturn"],                                   "Techno",             5, COLOR_PURPLE),
        # Acid House / Acid Techno
        (["808 state", "plastikman", "acid house",
          "acid track", "303", "acid techno"],          "Acid",               4, COLOR_PURPLE),
        # Jungle — pre-D&B breakbeat jungle (before D&B rule)
        (["roni size", "ltj bukem", "goldie - ",
          "jungle", "liquid funk jungle"],              "Jungle",             4, COLOR_ORANGE),
        # Drum & Bass
        (["drum and bass", "drum & bass", "dnb", "d&b",
          "liquid funk", "logistics", "chase & status",
          "andy c", "wilkinson", "high contrast",
          "pendulum", "netsky", "sub focus"],           "Drum & Bass",        4, COLOR_ORANGE),
        # Breakbeat / Big Beat  [v5.0 FIX: "the_chemical_brothers" underscore variant]
        (["crystal method", "the prodigy", "chemical brothers",
          "the_chemical_brothers",
          "fatboy slim", "basement jaxx", "2 many djs",
          "breakbeat", "breaks", "big beat",
          # v6.0
          "funk hunters", "the allergies", "deekline",
          "a-trak", "serial killaz"],                   "Breakbeat",          4, COLOR_ORANGE),
        # Bass / Dubstep
        (["dubstep", "skrillex", "zomboy", "brostep",
          "excision", "datsik", "flux pavilion",
          # v6.0
          "nghtmre", "slander", "hucci", "jaykode",
          "matrix & futurebound", "acheless",
          "cirqular", "subb theory",
          "with you friends", "deekapz"],               "Bass/Dubstep",       5, COLOR_RED),
        # Glitch Hop
        (["glitch mob", "glitch hop", "glitch",
          "the glitch mob", "amon tobin", "edIT"],      "Glitch Hop",         4, COLOR_PURPLE),
        # ── TRIP-HOP / CHILL CLUSTER ────────────────────────────────────────
        # Trip-Hop / Electronic Soul  [v5.0 NEW]
        # Moby (~4), Portishead, Massive Attack confirmed unclassified in Electronic
        (["moby", "portishead", "massive attack",
          "tricky - ", "unkle - ", "archive - ",
          "zero 7", "lemon jelly",
          "trip-hop", "triphop", "trip hop"],           "Trip-Hop",           3, COLOR_PURPLE),
        # Downtempo / Chillout  [v5.0 EXPANDED: +Khruangbin, Kerala Dust, Nils Frahm]
        (["bonobo", "four tet", "four-tet", "caribou",
          "o'flynn", "oflynn", "parallel", "kiasmos",
          "boards of canada", "tycho", "emancipator",
          "khruangbin", "kerala dust",
          "nils frahm", "the japanese house",
          "downtempo", "chillout",
          # v6.0
          "yosi horikawa", "blundetto", "debruit",
          "lo wolf", "siren tourist",
          "umbilical moonrise", "cardiac half",
          "the incredible string band"],                "Downtempo/Chillout", 2, COLOR_AQUA),
        (["ambient", "atmospheric", "meditation"],      "Ambient",            1, COLOR_AQUA),
        (["lo-fi", "lofi", "chillstep",
          # v6.0
          "autumn glow", "bliss looper", "peace sine",
          "orenda", "hohm - ", "claraty",
          "intr0beatz", "sleepy fish"],                 "Lo-Fi/Chill",        2, COLOR_AQUA),
        # ── HOUSE-ADJACENT ─────────────────────────────────────────────────
        # French House / Filter House
        (["daft punk", "french house", "filter house",
          "bob sinclar", "cassius", "modjo",
          "etienne de crecy", "alan braxe", "stardust",
          # v6.0
          "mr oizo", "supermen lovers", "bellaire",
          "tony shades", "deborah aime la bagarre"],    "French House",       4, COLOR_RED),
        # Organic Bass / Flow Arts  [v5.0 NEW]
        # Polish Ambassador, Equanimous, Mfinity confirmed unclassified
        (["polish ambassador", "equanimous", "mfinity",
          "kalya scintilla", "east forest",
          "opiou", "bird of prey - ",
          "organic bass", "flow arts",
          "psybient"],                                  "Organic Bass",       3, COLOR_AQUA),
        # ── INDIE / ALT ELECTRONIC ──────────────────────────────────────────
        # Indie Electronic  [v5.0 EXPANDED: +The XX, Beach House, Cigs After Sex]
        (["indie", "shoegaze", "dream pop", "art school",
          "arthur russell", "glass animals", "sylvan esso",
          "golden vessel", "whethan",
          "the xx", "beach house",
          "cigarettes after sex", "alvvays",
          "wild nothing", "memoryhouse",
          # v6.0
          "roosevelt", "mura masa", "baynk",
          "salt cathedral", "møme", "fakear",
          "benji lewis", "joey pecoraro", "j ember",
          "koreless", "anna meredith", "kingdom -",
          "hundred waters", "idris dau", "frederic robinson",
          "shadow age", "rikslyd", "adi ulmansky",
          "fasme", "mou5zyzz", "tentendo"],             "Indie Electronic",   3, COLOR_BLUE),
        # Italo / Synth — catch-all for remaining Italo variants
        (["italo", "synth-pop", "new wave"],            "Italo/Synth",        3, COLOR_YELLOW),
        # ── TECH HOUSE  [v6.0 NEW] ──────────────────────────────────────────
        (["camelphat", "piero pirupa", "pleasurekraft",
          "lo'99", "taiki nulight", "mochakk",
          "tony romera", "moreno pezzolato", "osunlade",
          "riton", "countach", "pilocka krach",
          "etienne jaumet", "tom wax", "harrison bdp"],  "Tech House",        4, COLOR_PURPLE),
        # ── MELODIC HOUSE  [v6.0 NEW] ───────────────────────────────────────
        (["jan blomqvist", "oliver schories", "pachanga boys",
          "carlita", "maribou state", "rigopolar",
          "autoload"],                                   "Melodic House",     4, COLOR_PURPLE),
        # ── CLUB / DANCE POP  [v6.0 NEW] ────────────────────────────────────
        # Covers chart-dance floor-fillers: Eurodance, EDM-pop, party anthems
        (["alan walker", "@vize", "vize x",
          "alexandra stan", "aronchupa",
          "enur feat", "guru josh", "laurent wolf",
          "loud luxury", "r.i.o.", "wamdue project",
          "sugababes", "sean kingston", "fergie ft",
          "die atzen", "deichkind", "dr' alban", "dr alban",
          "snollebolleke", "twocolors", "rob & chris",
          "rudenko", "ida corr", "ian carey", "coro feat",
          "passion fruit", "calabria 2007",
          "king of my castle", "get shaky",
          "numa numa", "omi feat",
          "alexander marcus",
          "oka - ", "bomel", "levyticus",
          "joost ", "don_t_stop"],                       "Club/Dance Pop",    4, COLOR_ORANGE),
        # Future Bass  [v5.0 EXPANDED: +Flume, Illenium, Big Wild, Medasin]
        (["future bass", "trap edm",
          "flume", "illenium - ",
          "big wild", "medasin",
          "melodic trap"],                              "Future Bass",        4, COLOR_ORANGE),
    ],

    # ── POP ──────────────────────────────────────────────────────────────────
    "Pop": [
        # Musical / Broadway
        (["hamilton", "broadway", "musical", "frozen", "lion king",
          "grease", "you'll be back", "i know him"],    "Musical/Show",       3, COLOR_YELLOW),
        # K-Pop — BTS / BLACKPINK generation
        (["bts", "blackpink", "psy - ", "twice - ",
          "stray kids", "exo - ", "k-pop", "kpop",
          "gangnam"],                                   "K-Pop",              4, COLOR_PINK),
        # 80s Pop — Michael Jackson, Madonna era
        (["george michael", "wham!", "michael jackson",
          "madonna", "culture club", "cyndi lauper",
          "duran duran", "a-ha", "human league",
          "eurythmics"],                                "80s Pop",            4, COLOR_YELLOW),
        # Indie Pop
        (["indie pop", "her's", "devendra", "kai straw",
          "madison ryann", "healy", "koosen", "d4vd"],  "Indie Pop",          3, COLOR_PINK),
        # 90s Pop
        (["spice", "nsync", "backstreet", "ace of base", "whigfield",
          "saturday night", "bye bye bye"],             "90s Pop",            4, COLOR_YELLOW),
        # ABBA / Classics
        (["abba", "gimme!", "take a chance"],           "ABBA/Classics",      4, COLOR_YELLOW),
        # Dance-Pop
        (["dance pop", "walk the moon", "maroon 5", "sugar"],
                                                        "Dance-Pop",          4, COLOR_ORANGE),
        # Acoustic Pop
        (["acoustic", "stripped", "folk pop"],          "Acoustic Pop",       2, COLOR_BLUE),
    ],

    # ── ROCK  (v3.0 — Death Metal, Hair Metal/Glam, Soft Rock added) ─────────
    "Rock": [
        # Nu-Metal — confirmed: Linkin Park ×6, Limp Bizkit ×5, Papa Roach ×3
        (["linkin park", "limp bizkit", "papa roach", "deftones",
          "disturbed", "mudvayne", "godsmack", "korn",
          "slipknot", "drowning pool", "static-x",
          "p.o.d", "nookie", "numb ", "in the end",
          "crawling", "hybrid theory"],                 "Nu-Metal",           5, COLOR_RED),
        # Thrash / Heavy Metal — confirmed: Metallica ×many
        (["metallica", "megadeth", "anthrax", "slayer",
          "pantera", "sepultura", "machine head",
          "thrash", "master of puppets", "battery",
          "one - ", "for whom the bell", "nothing else matters",
          "enter sandman"],                             "Thrash Metal",       5, COLOR_RED),
        # Death Metal
        (["cannibal corpse", "morbid angel", "obituary",
          "deicide", "suffocation", "possessed - ",
          "death metal"],                               "Death Metal",        5, COLOR_RED),
        # Hard Rock — classic heavy rock (pre-grunge)
        (["ac/dc", "acdc", "ozzy", "black sabbath", "deep purple",
          "led zeppelin", "rainbow", "dio", "whitesnake",
          "guns n roses", "guns n' roses", "motley crue",
          "van halen", "boston - ", "scorpions",
          "kiss - ", "queensrÿche", "queensryche",
          "hard rock"],                                 "Hard Rock",          4, COLOR_ORANGE),
        # Hair Metal / Glam Rock — 80s glamour
        (["poison - ", "warrant - ", "cinderella - ",
          "ratt - ", "winger - ", "skid row",
          "trixter", "quireboys", "hair metal",
          "glam metal", "glam rock"],                   "Hair Metal/Glam",    4, COLOR_ORANGE),
        # Soft Rock — AOR / easy-listening rock
        (["eagles - ", "journey - ", "billy joel",
          "elton john", "christopher cross",
          "hall & oates", "hall and oates",
          "toto - ", "styx - ", "reo speedwagon",
          "america - ", "fleetwood mac - you make",
          "soft rock"],                                 "Soft Rock",          3, COLOR_BLUE),
        # Alternative Metal — Tool, Nine Inch Nails, RATM
        (["a perfect circle", "nine inch nails", "nin - ",
          "tool - ", "rage against", "ratm", "audioslave",
          "chris cornell", "soundgarden - ",
          "alice in chains - ",
          "alt metal", "alternative metal"],            "Alternative Metal",  4, COLOR_PURPLE),
        # Post-Grunge — confirmed: 3 Doors Down ×4
        (["3 doors down", "nickelback", "creed - ",
          "puddle of mudd", "staind", "default - ",
          "hinder", "lifehouse", "matchbox twenty",
          "matchbox 20", "post-grunge"],                "Post-Grunge",        4, COLOR_BLUE),
        # Grunge — confirmed: Nirvana, Pearl Jam, Soundgarden, Alice in Chains
        (["nirvana", "pearl jam", "soundgarden",
          "alice in chains",
          "mudhoney", "screaming trees", "stone temple pilots",
          "grunge"],                                    "Grunge",             4, COLOR_PURPLE),
        # Punk / Pop-Punk / Hardcore — confirmed: Green Day, blink-182
        (["green day", "blink-182", "blink 182",
          "the offspring", "rancid", "bad religion",
          "social distortion", "nofx", "descendents",
          "sex pistols", "the clash", "ramones",
          "punk", "hardcore", "emo",
          "television personalities"],                  "Punk",               4, COLOR_ORANGE),
        # Indie Rock — Arctic Monkeys, Pixies, Tame Impala etc.
        (["arctic monkeys", "pixies", "belle and sebastian",
          "belle & sebastian", "built to spill",
          "dinosaur jr", "animal collective", "white stripes",
          "the strokes", "vampire weekend", "tame impala",
          "no age", "marine girls", "raincoats",
          "modest mouse", "pavement", "guided by voices"],
                                                        "Indie Rock",         3, COLOR_BLUE),
        # New Wave / Post-Punk
        (["depeche mode", "new order", "joy division",
          "the chameleons", "the cure", "siouxsie",
          "bauhaus", "the smiths", "the fall",
          "talking heads", "blondie", "post-punk",
          "new wave"],                                  "New Wave/Post-Punk", 3, COLOR_BLUE),
        # Psychedelic / Shoegaze
        (["nick cave", "shoegaze", "my bloody valentine",
          "slowdive", "ride - ", "mazzy star",
          "spacemen 3", "spiritualized", "the brian jonestown",
          "psychedelic", "13th floor elevators"],       "Psychedelic",        3, COLOR_PURPLE),
        # Folk-Rock / Acoustic
        (["folk rock", "acoustic", "sufjan stevens",
          "iron & wine", "fleet foxes", "bon iver",
          "joshua tree", "simon & garfunkel"],          "Folk-Rock",          2, COLOR_BLUE),
        # Classic Rock — massively expanded artist list
        (["aerosmith", "bon jovi - ", "joan jett",
          "kate bush", "fleetwood mac", "lynyrd skynyrd",
          "ram jam", "creedence", "r.e.m - ", "rem - ",
          "powderfinger", "the who", "the rolling stones",
          "rolling stones", "jimi hendrix",
          "queen - ", "cream - ", "the police",
          "dire straits", "u2 - ", "bruce springsteen",
          "tom petty", "zz top", "steve miller",
          "heart - ", "bryan adams",
          "david bowie", "sting - ",
          "the beatles", "beatles",
          "the doors", "doors - ",
          "bob dylan - ", "neil young",
          "eric clapton", "peter gabriel",
          "classic rock"],                              "Classic Rock",       3, COLOR_ORANGE),
        # Modern Rock — catch-all for recent guitar rock
        (["ghost - ", "queens of the stone age", "qotsa",
          "foo fighters", "royal blood", "wolfmother",
          "modern rock"],                               "Modern Rock",        3, COLOR_BLUE),
    ],

    # ── LATIN ────────────────────────────────────────────────────────────────
    "Latin": [
        # Reggaeton — floor-filler; confirmed artists
        (["reggaeton", "bad bunny", "j balvin", "daddy yankee", "maluma",
          "farruko", "don omar", "ozuna", "nicky jam", "te bote",
          "rompe", "pepas", "veneno", "movimiento"],    "Reggaeton",          5, COLOR_RED),
        # Bachata — romantic guitar; Romeo Santos era
        (["romeo santos", "prince royce", "aventura",
          "juan luis guerra", "bachata"],               "Bachata",            4, COLOR_PINK),
        # Latin Trap — urban latino
        (["anuel aa", "bryant myers", "jhay cortez",
          "latin trap", "urbano", "trap latino"],       "Latin Trap",         5, COLOR_RED),
        # Dembow — Dominican club music
        (["el alfa", "dembow", "tekashi"],              "Dembow",             5, COLOR_RED),
        # Cumbia / Vallenato
        (["cumbia", "vallenato"],                       "Cumbia",             4, COLOR_ORANGE),
        # Salsa / Merengue
        (["salsa", "merengue"],                         "Salsa",              4, COLOR_YELLOW),
        # Latin Folk / Bossa
        (["bossa nova", "samba", "brasil", "brazil",
          "mercedes sosa", "gracias a la vida"],        "Latin Folk/Bossa",   2, COLOR_AQUA),
        # Latin Pop
        (["jennifer lopez", "ricky martin", "shakira", "selena",
          "j. lo", "livin la", "la bamba"],             "Latin Pop",          4, COLOR_PINK),
        # Tango / Flamenco
        (["flamenco", "tango", "chingon", "malagueña", "malaguen"],
                                                        "Tango/Flamenco",     3, COLOR_YELLOW),
        # Soca / Caribbean
        (["soca", "machel", "superblue", "caribbean"],  "Soca",               4, COLOR_ORANGE),
    ],

    # ── CLASSICS  (v3.0 — Blues added) ──────────────────────────────────────
    "Classics": [
        # Disco — confirmed: Donna Summer, Earth Wind & Fire
        (["disco", "donna summer", "earth wind", "kc and",
          "gloria gaynor", "village people", "bee gees",
          "hot stuff", "somebody dance", "yowsah"],     "Disco",              5, COLOR_ORANGE),
        # Soul / Motown — James Brown, Aretha, Marvin Gaye
        (["soul", "motown", "aretha", "james brown", "marvin gaye",
          "otis redding", "nina simone", "bill withers",
          "al green", "you got the funk"],              "Soul/Motown",        4, COLOR_ORANGE),
        # Blues — BB King, Stevie Ray Vaughan, classics
        (["bb king", "muddy waters", "robert johnson",
          "howlin wolf", "howlin' wolf", "buddy guy",
          "john lee hooker", "stevie ray vaughan",
          "blues"],                                     "Blues",              3, COLOR_BLUE),
        # 50s — confirmed: Elvis, Chuck Berry
        (["1950", "1951", "1952", "1953", "1954", "1955",
          "1956", "1957", "1958", "1959",
          "elvis", "chuck berry", "little richard",
          "buddy holly", "bo diddley", "fats domino",
          "bill haley", "rock around the clock"],       "50s",                3, COLOR_YELLOW),
        (["1960", "1961", "1962", "1963", "1964", "1965", "1966", "1967",
          "1968", "1969", "love - ", "chad & jeremy",
          "kinks", "zombies", "beach boys", "ronettes",
          "bob dylan"],                                 "60s",                3, COLOR_YELLOW),
        (["1970", "1971", "1972", "1973", "1974", "1975", "1976", "1977",
          "1978", "1979", "fleetwood mac - ", "lynyrd", "boney m",
          "dexys", "earth wind & fire"],               "70s/Funk",           3, COLOR_YELLOW),
        (["1980", "1981", "1982", "1983", "1984", "1985", "1986", "1987",
          "1988", "1989",
          "berlin - ", "bronski beat", "orchestral manoeuv"],
                                                        "80s",                3, COLOR_BLUE),
        (["1990", "1991", "1992", "1993", "1994", "1995", "1996", "1997",
          "1998", "1999",
          "nirvana - ", "oasis - ", "blur - ", "pulp - ",
          "suede - "],                                  "90s",                3, COLOR_BLUE),
        # Film OST / Soundtrack
        (["soundtrack", " ost", "ost ", "film score", "pulp fiction",
          "hateful eight", "once upon a time", "green hornet",
          "lo chiamavano", "son of a lovin"],           "Film OST",           3, COLOR_PURPLE),
        # Jazz / Swing / Big Band
        (["swing", "jazz", "big band", "sinatra", "dean martin", "frank"],
                                                        "Jazz/Swing",         3, COLOR_BLUE),
        # Camp / Drag
        (["rupaul", "glamazon"],                        "Camp/Drag",          4, COLOR_PINK),
    ],

    # ── WORLD ────────────────────────────────────────────────────────────────
    "World": [
        # Reggae — Bob Marley, Sean Paul, Ska
        (["reggae", "bob marley", "ska", "rocksteady", "toots",
          "slickers", "shaggy", "boombastic", "misty morning",
          "is this love", "sean paul"],                 "Reggae",             3, COLOR_GREEN),
        # Dub
        (["dub", "o.b.f", "obf", "pressure drop"],     "Dub",                2, COLOR_AQUA),
        # Amapiano — log-drum basslines, 110–115 BPM, South African
        # (Google Doc 2025 micro-genres; also in Electronic but DJ crossover in World)
        (["amapiano", "log drum", "log-drum",
          "kabza de small", "dj maphorisa", "focalistic",
          "ami faku", "sun-el musician"],               "Amapiano",           4, COLOR_ORANGE),
        # Afrobeats (modern — Burna Boy era)
        (["burna boy", "wizkid", "davido", "mr eazi",
          "adekunle gold", "afrobeats"],                "Afrobeats",          4, COLOR_ORANGE),
        # Afrobeat/Funk (classic — Fela Kuti era)
        (["afrobeat", "africa", "voilaaa", "mory kante", "yeke yeke",
          "lat teens", "fela kuti", "tony allen"],      "Afrobeat/Funk",      4, COLOR_ORANGE),
        # French Pop — Stromae, Indila generation
        (["stromae", "christine and the queens",
          "indila", "yelle", "carla bruni",
          "french pop"],                                "French Pop",         3, COLOR_PINK),
        # Greek / Turkish — Mediterranean pop
        (["greek", "turkish", "sertab erener",
          "tarkan", "antique"],                         "Greek/Turkish",      3, COLOR_YELLOW),
        # Ecstatic Dance — festival flow arts scene
        (["ecstatic", "equanimous", "bird tribe", "akriza", "activation",
          "banyan", "jpattersson", "vision dojo", "human experience",
          "julietta", "ancient future"],                "Ecstatic Dance",     4, COLOR_AQUA),
        # Yoga / Sacred
        (["yoga", "meditation", "mantra", "kirtan", "shakti", "shiva moon",
          "prem joshua", "tapatham", "dj drez"],        "Yoga/Sacred",        2, COLOR_AQUA),
        # Balkan / Klezmer
        (["klezmer", "balkan", "fanfare ciocarlia", "balkan beat box",
          "zventa sventana", "bulgarian"],              "Balkan/Klezmer",     4, COLOR_ORANGE),
        # Arabic / Middle East
        (["arabic", "middle east", "omar souleyman", "warni warni",
          "enta omri", "drumspyder", "saiidi", "شارموفرز"],
                                                        "Arabic/Middle East", 3, COLOR_YELLOW),
        # Indian / Bollywood
        (["indian", "tabla", "sitar", "bollywood", "tapatham"],
                                                        "Indian",             3, COLOR_YELLOW),
        # Global Bass
        (["thornato", "captain planet", "global bass"], "Global Bass",        4, COLOR_ORANGE),
    ],

    # ── TOOLS & FX ───────────────────────────────────────────────────────────
    "Tools & FX": [
        (["drum", "kick", "snare", "hat", "clap", "cymbal", "dhol",
          "djembe", "clave", "cowbell", "agogo", "dholak"],
                                                        "Drum/Percussion",    1, COLOR_GREEN),
        (["bass", "808", "sub "],                       "Bass/808",           1, COLOR_GREEN),
        (["fx", "sfx", "effect", "sweep", "riser"],     "FX",                 1, COLOR_GREEN),
        (["loop", "sample", "intro", "transition"],     "Loop/Transition",    1, COLOR_GREEN),
    ],

    # ── REMIXES  (v2.0 — Electro Swing Remix leads) ──────────────────────────
    "Remixes": [
        # Electro Swing Remix — confirmed: Club Des Belugas ×2
        (["club des belugas", "electro swing", "electroswing",
          "sim gretina", "parov stelar", "caravan palace",
          "caro emerald"],                              "Electro Swing Remix",4, COLOR_YELLOW),
        (["mashup", "mash-up", " vs ", " x "],          "Mashup",             4, COLOR_ORANGE),
        (["bootleg", "unofficial"],                     "Bootleg",            3, COLOR_ORANGE),
        (["regrooved", "rework", "re-edit", "re edit"], "Rework",             3, COLOR_YELLOW),
        (["cover"],                                     "Cover",              3, COLOR_PINK),
        (["remix", "rmx"],                              "Remix",              3, COLOR_ORANGE),
    ],

    "INBOX": [],
}

# ── Energy keyword modifiers ───────────────────────────────────────────────────
ENERGY_DOWN = ["intro ", " intro", "outro ", " outro", "ambient", "yoga",
               "meditation", "stripped", "acoustic version", "lo-fi", "lofi",
               "downtempo", "slow jam", "chill"]
ENERGY_UP   = ["banger", "club mix", "festival", "peak", "hard mix",
               "banging", "floor filler", "rave", "dance mix"]

# ── Energy tier mapping (v5.0) ─────────────────────────────────────────────
# Applied AFTER ENERGY_DOWN / ENERGY_UP star modifiers. Stars → energy label.
STAR_TO_ENERGY: dict[int, str] = {
    5: "PEAK",
    4: "HIGH",
    3: "MID",
    2: "LOW",
    1: "LOW",
}

# ──────────────────────────────────────────────────────────────────────────────
# YEAR ERA FALLBACK  (v2.0 — fires when no keyword rule matches)
# Maps genre → {decade_int: era_label}
# decade key = (year // 10) * 10  →  1993 → 1990, 2003 → 2000
# ──────────────────────────────────────────────────────────────────────────────
YEAR_ERA_FALLBACK = {
    "Rock": {
        1950: "50s Rock", 1960: "60s Rock", 1970: "70s Rock",
        1980: "80s Rock", 1990: "90s Rock", 2000: "00s Rock",
        2010: "10s Rock", 2020: "20s Rock",
    },
    "Hip-Hop & R&B": {
        1980: "80s Hip-Hop", 1990: "Golden Age Hip-Hop",
        2000: "00s Hip-Hop", 2010: "10s Hip-Hop", 2020: "New School",
    },
    "Electronic": {
        1980: "80s Electronic", 1990: "90s Electronic",
        2000: "00s Electronic", 2010: "10s Electronic", 2020: "20s Electronic",
    },
    "House & Dance": {
        1980: "80s House", 1990: "90s House",
        2000: "00s House", 2010: "10s House", 2020: "20s House",
    },
    "Pop": {
        1960: "60s Pop", 1970: "70s Pop", 1980: "80s Pop",
        1990: "90s Pop", 2000: "00s Pop", 2010: "10s Pop", 2020: "20s Pop",
    },
    "Latin": {
        1980: "80s Latin", 1990: "90s Latin",
        2000: "00s Latin", 2010: "10s Latin", 2020: "20s Latin",
    },
    "Classics": {
        1920: "Jazz Age", 1930: "Swing Era", 1940: "40s",
        1950: "50s", 1960: "60s", 1970: "70s/Funk",
        1980: "80s", 1990: "90s", 2000: "00s Classic",
    },
    "Israeli": {
        1990: "90s Israeli", 2000: "00s Israeli",
        2010: "10s Israeli", 2020: "20s Israeli",
    },
}


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s)


def get_genre_config(folder_name: str):
    """Map a folder name to (genre_label, default_stars, default_color, default_subgenre)."""
    fn_lower = folder_name.lower()
    for prefix, genre, stars, color, subgenre in GENRE_CONFIG:
        if fn_lower.startswith(prefix):
            return genre, stars, color, subgenre
    return "Unknown", 3, COLOR_NONE, "Unknown"


def classify_track(filename: str, genre: str):
    """
    Determine (subgenre, stars, color, energy) for a track.

    Priority order:
      1. Keyword rules in SUBGENRE_RULES  (first match wins)
      2. Year-based era fallback in YEAR_ERA_FALLBACK
      3. Genre default from GENRE_CONFIG

    Returns: (subgenre_str, stars_int, color_int, energy_str)
    Energy is derived from final stars after ENERGY_DOWN / ENERGY_UP modifiers.
    """
    name_lower = nfc(filename.lower())
    rules = SUBGENRE_RULES.get(genre, [])

    subgenre, stars, color = None, None, None

    # ── Step 1: keyword rules ─────────────────────────────────────────────
    for keywords, sg, st, co in rules:
        if any(kw in name_lower for kw in keywords):
            subgenre, stars, color = sg, st, co
            break

    # ── Step 2: year-based era fallback ──────────────────────────────────
    if subgenre is None:
        year_str = extract_year(filename)
        if year_str:
            era_map = YEAR_ERA_FALLBACK.get(genre, {})
            dk = (int(year_str) // 10) * 10
            if dk in era_map:
                for prefix, gl, ds, dc, dsg in GENRE_CONFIG:
                    if gl == genre:
                        subgenre, stars, color = era_map[dk], ds, dc
                        break

    # ── Step 3: genre defaults ────────────────────────────────────────────
    if subgenre is None:
        for prefix, gl, ds, dc, dsg in GENRE_CONFIG:
            if gl == genre:
                subgenre, stars, color = dsg, ds, dc
                break
    if subgenre is None:
        subgenre, stars, color = "Unknown", 3, COLOR_NONE

    # ── Energy modifiers (clamp 1–5) ──────────────────────────────────────
    if any(kw in name_lower for kw in ENERGY_DOWN):
        stars = max(1, stars - 1)
    if any(kw in name_lower for kw in ENERGY_UP):
        stars = min(5, stars + 1)

    # Tools always = 1 star
    if genre == "Tools & FX":
        stars = 1

    # ── Energy tier (derived from final stars) ────────────────────────────
    energy = STAR_TO_ENERGY.get(int(stars), "MID")

    return subgenre, int(stars), int(color), energy


def extract_year(filename: str) -> str:
    """Extract a plausible year (1950–2029) from the filename string."""
    m = YEAR_RE.search(filename)
    return m.group(1) if m else ""


def parse_artist_title(filename: str):
    """
    Best-effort parse of 'Artist - Title' from filename.
    Returns (artist, title).
    """
    stem = Path(filename).stem
    # Try 'Artist - Title'
    if " - " in stem:
        parts = stem.split(" - ", 1)
        return parts[0].strip(), parts[1].strip()
    # Try 'NN Artist - Title' (track number prefix)
    m = re.match(r'^\d+[\s\-_.]+(.+)', stem)
    if m:
        inner = m.group(1)
        if " - " in inner:
            parts = inner.split(" - ", 1)
            return parts[0].strip(), parts[1].strip()
        return "", inner.strip()
    return "", stem.strip()


def file_to_mac_location(vm_path: Path) -> str:
    """Convert a VM path to a Rekordbox-compatible file:// URL for the Mac."""
    rel = vm_path.relative_to(VM_DJMUSIC)
    mac_path = MAC_DJMUSIC + "/" + str(rel).replace("\\", "/")
    return "file://" + urllib.parse.quote(mac_path, safe="/:")


# ──────────────────────────────────────────────────────────────────────────────
# SCANNER — build list of track dicts
# ──────────────────────────────────────────────────────────────────────────────
def scan_library() -> list[dict]:
    tracks = []
    track_id = 1

    for folder in sorted(VM_DJMUSIC.iterdir()):
        if not folder.is_dir():
            continue
        genre, def_stars, def_color, def_subgenre = get_genre_config(folder.name)

        for f in sorted(folder.iterdir()):
            if f.suffix.lower() not in AUDIO_EXTS:
                continue

            artist, title = parse_artist_title(f.name)
            subgenre, stars, color, energy = classify_track(f.name, genre)
            year = extract_year(f.name)

            tracks.append({
                "id":       track_id,
                "path":     f,
                "folder":   folder.name,
                "filename": f.name,
                "artist":   artist,
                "title":    title,
                "genre":    genre,
                "subgenre": subgenre,
                "energy":   energy,
                "stars":    stars,
                "color":    color,
                "color_name": COLOR_NAMES[color],
                "year":     year,
                "ext":      f.suffix.lower(),
                "location": file_to_mac_location(f),
                "kind":     EXT_TO_KIND.get(f.suffix.lower(), "Audio File"),
            })
            track_id += 1

    return tracks


# ──────────────────────────────────────────────────────────────────────────────
# REKORDBOX XML GENERATOR
# ──────────────────────────────────────────────────────────────────────────────
def write_rekordbox_xml(tracks: list[dict]):
    """Generate rekordbox_import.xml for Rekordbox Library import."""
    root = ET.Element("DJ_PLAYLISTS", Version="1")
    ET.SubElement(root, "PRODUCT",
                  Name="rekordbox", Version="6.0.0", Company="AlphaTheta")

    collection = ET.SubElement(root, "COLLECTION", Entries=str(len(tracks)))
    for t in tracks:
        attribs = {
            "TrackID":     str(t["id"]),
            "Name":        t["title"] or t["filename"],
            "Artist":      t["artist"],
            "Genre":       t["genre"],
            "Kind":        t["kind"],
            "TotalTime":   "0",
            "DiscNumber":  "0",
            "TrackNumber": "0",
            "Year":        t["year"],
            "Bpm":         "0.00",
            "DateAdded":   TODAY,
            "BitRate":     "0",
            "SampleRate":  "44100",
            "Comments":    f"{t['subgenre']} | {t['energy']}",
            "Rating":      str(STARS_TO_RATING[t["stars"]]),
            "Location":    t["location"],
            "Remixer":     "",
            "Tonality":    "",
            "Label":       "",
            "Mix":         "",
            "Colour":      str(t["color"]),
        }
        ET.SubElement(collection, "TRACK", **attribs)

    # Build one playlist per genre folder + a master "All Tracks" playlist
    playlists_root = ET.SubElement(root, "PLAYLISTS")
    root_node = ET.SubElement(playlists_root, "NODE",
                              Type="0", Name="ROOT", Count="2")

    # All Tracks playlist
    all_node = ET.SubElement(root_node, "NODE",
                             Name="All Tracks", Type="1", KeyType="0",
                             Entries=str(len(tracks)))
    for t in tracks:
        ET.SubElement(all_node, "TRACK", Key=str(t["id"]))

    # Per-genre playlists
    genres_node = ET.SubElement(root_node, "NODE",
                                Type="0", Name="By Genre", Count="0")
    genre_groups: dict[str, list] = {}
    for t in tracks:
        genre_groups.setdefault(t["genre"], []).append(t)

    for genre_label, g_tracks in sorted(genre_groups.items()):
        gnode = ET.SubElement(genres_node, "NODE",
                              Name=genre_label, Type="1", KeyType="0",
                              Entries=str(len(g_tracks)))
        for t in g_tracks:
            ET.SubElement(gnode, "TRACK", Key=str(t["id"]))

    # Pretty-print
    ET.indent(root, space="  ")
    tree = ET.ElementTree(root)
    tree.write(str(XML_OUT), encoding="utf-8", xml_declaration=True)
    print(f"✅  Wrote {len(tracks)} tracks to {XML_OUT.name}")


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATOR
# ──────────────────────────────────────────────────────────────────────────────
def write_xlsx(tracks: list[dict]):
    """Generate library_tags.xlsx — master spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "All Tracks"

    # ── Column headers ─────────────────────────────────────────────────────
    headers = ["#", "Folder", "Artist", "Title", "Genre", "Sub-Genre",
               "Energy", "Stars ★", "Color", "Year", "Filename"]
    col_widths = [5, 22, 30, 45, 18, 22, 8, 9, 10, 6, 60]

    # Header style
    hdr_fill  = PatternFill("solid", fgColor="1A1A2E")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    thin_side = Side(style="thin", color="D0D0D0")
    thin_bdr  = Border(bottom=thin_side)

    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[1].height = 22

    # ── Color map for color-coded rows ────────────────────────────────────
    FILL_MAP = {
        COLOR_RED:    "FFD5D5",
        COLOR_ORANGE: "FFE8C8",
        COLOR_YELLOW: "FFFACC",
        COLOR_PINK:   "FFE0F0",
        COLOR_BLUE:   "D5E8FF",
        COLOR_PURPLE: "E8D5FF",
        COLOR_AQUA:   "D5F5F5",
        COLOR_GREEN:  "D5F5D5",
        COLOR_NONE:   "F0F0F0",
    }
    star_str = {1:"★", 2:"★★", 3:"★★★", 4:"★★★★", 5:"★★★★★"}

    for row_i, t in enumerate(tracks, start=2):
        row_data = [
            t["id"],
            t["folder"],
            t["artist"],
            t["title"] or t["filename"],
            t["genre"],
            t["subgenre"],
            t["energy"],
            star_str.get(t["stars"], str(t["stars"])),
            t["color_name"],
            t["year"],
            t["filename"],
        ]
        fill = PatternFill("solid", fgColor=FILL_MAP.get(t["color"], "FFFFFF"))
        for col_i, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_bdr

        ws.row_dimensions[row_i].height = 15

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    summary_headers = ["Genre", "Tracks", "Sub-genres Used",
                       "Avg Stars ★", "Colors Used"]
    for col_i, h in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col_i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A1A2E")
        cell.alignment = Alignment(horizontal="center")

    genre_summary: dict[str, dict] = {}
    for t in tracks:
        g = t["genre"]
        if g not in genre_summary:
            genre_summary[g] = {"count": 0, "stars": [], "subgenres": set(), "colors": set()}
        genre_summary[g]["count"] += 1
        genre_summary[g]["stars"].append(t["stars"])
        genre_summary[g]["subgenres"].add(t["subgenre"])
        genre_summary[g]["colors"].add(t["color_name"])

    for row_i, (genre, data) in enumerate(sorted(genre_summary.items()), start=2):
        avg_stars = sum(data["stars"]) / len(data["stars"])
        ws2.cell(row=row_i, column=1, value=genre).font = Font(bold=True)
        ws2.cell(row=row_i, column=2, value=data["count"])
        ws2.cell(row=row_i, column=3, value=", ".join(sorted(data["subgenres"])))
        ws2.cell(row=row_i, column=4, value=f"{avg_stars:.1f}")
        ws2.cell(row=row_i, column=5, value=", ".join(sorted(data["colors"])))

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["C"].width = 80
    ws2.column_dimensions["E"].width = 40

    wb.save(str(XLSX_OUT))
    print(f"✅  Wrote {len(tracks)} tracks to {XLSX_OUT.name}")


# ──────────────────────────────────────────────────────────────────────────────
# ID3 TAG WRITER
# ──────────────────────────────────────────────────────────────────────────────
def write_id3_tags(tracks: list[dict]):
    """Write Genre, Comment (sub-genre | energy), Year ID3 tags to audio files."""
    from mutagen.mp3  import MP3
    from mutagen.id3  import ID3, TIT2, TPE1, TCON, COMM, TDRC, ID3NoHeaderError
    from mutagen.flac import FLAC
    from mutagen.mp4  import MP4
    from mutagen.wave import WAVE

    written = 0
    errors  = 0

    for t in tracks:
        path        = t["path"]
        ext         = t["ext"]
        comment_val = f"{t['subgenre']} | {t['energy']}"
        try:
            if ext == ".mp3":
                try:
                    tags = ID3(str(path))
                except ID3NoHeaderError:
                    tags = ID3()
                tags.add(TCON(encoding=3, text=[t["genre"]]))
                tags.add(COMM(encoding=3, lang="eng", desc="", text=[comment_val]))
                if t["year"]:
                    tags.add(TDRC(encoding=3, text=[t["year"]]))
                tags.save(str(path))
                written += 1

            elif ext == ".flac":
                audio = FLAC(str(path))
                audio["genre"]   = t["genre"]
                audio["comment"] = comment_val
                if t["year"]:
                    audio["date"] = t["year"]
                audio.save()
                written += 1

            elif ext == ".m4a":
                audio = MP4(str(path))
                audio["\xa9gen"] = [t["genre"]]
                audio["\xa9cmt"] = [comment_val]
                if t["year"]:
                    audio["\xa9day"] = [t["year"]]
                audio.save()
                written += 1

            elif ext in (".wav", ".aif", ".aiff"):
                # WAV ID3 via mutagen
                try:
                    from mutagen.id3 import ID3
                    tags = ID3(str(path))
                except Exception:
                    tags = ID3()
                tags.add(TCON(encoding=3, text=[t["genre"]]))
                tags.add(COMM(encoding=3, lang="eng", desc="", text=[comment_val]))
                if t["year"]:
                    tags.add(TDRC(encoding=3, text=[t["year"]]))
                tags.save(str(path))
                written += 1

        except Exception as e:
            print(f"  ⚠️  Could not tag {path.name}: {e}", file=sys.stderr)
            errors += 1

    print(f"✅  Tagged {written} files  ({errors} errors)")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="tag_library.py — DJ library tagger v5.0"
    )
    parser.add_argument("--write-tags", action="store_true",
                        help="Write ID3 Genre/Comment/Year tags to audio files")
    parser.add_argument("--preview", action="store_true",
                        help="Print first 30 tracks to terminal, no files written")
    args = parser.parse_args()

    print("🎵  Scanning DJ_MUSIC library …")
    tracks = scan_library()
    print(f"    Found {len(tracks):,} audio files\n")

    if not tracks:
        print("❌  No audio files found. Check VM_DJMUSIC path.")
        sys.exit(1)

    if args.preview:
        # Print a pretty table of the first 30 tracks
        print(f"{'ID':>4}  {'Genre':<18}  {'Sub-Genre':<25}  {'★':>1}  {'Color':<8}  {'Yr':<4}  Filename")
        print("─" * 115)
        for t in tracks[:30]:
            print(f"{t['id']:>4}  {t['genre']:<18}  {t['subgenre']:<25}  "
                  f"{t['stars']}  {t['color_name']:<8}  {t['year']:<4}  "
                  f"{t['filename'][:60]}")
        print(f"\n    … {len(tracks) - 30} more tracks. Run without --preview to generate outputs.")
        return

    # Print genre summary
    genre_counts: dict[str, int] = {}
    for t in tracks:
        genre_counts[t["genre"]] = genre_counts.get(t["genre"], 0) + 1
    print("📊  Genre breakdown:")
    for g, n in sorted(genre_counts.items(), key=lambda x: -x[1]):
        print(f"    {g:<20} {n:>4} tracks")
    print()

    # Print sub-genre breakdown
    subgenre_counts: dict[str, dict[str, int]] = {}
    for t in tracks:
        g, sg = t["genre"], t["subgenre"]
        subgenre_counts.setdefault(g, {})
        subgenre_counts[g][sg] = subgenre_counts[g].get(sg, 0) + 1
    print("📊  Sub-genre breakdown:")
    for g, sg_counts in sorted(subgenre_counts.items()):
        genre_total = sum(sg_counts.values())
        print(f"\n  {g} ({genre_total}):")
        for sg, n in sorted(sg_counts.items(), key=lambda x: -x[1]):
            print(f"    {sg:<35} {n:>4}")
    print()

    write_rekordbox_xml(tracks)
    write_xlsx(tracks)

    if args.write_tags:
        print("\n🏷️   Writing ID3 tags to audio files …")
        write_id3_tags(tracks)

    print(f"""
╔══════════════════════════════════════════════════════════════════╗
║  ✅  All outputs ready in ~/Music/                               ║
║                                                                  ║
║  1. rekordbox_import.xml  →  Rekordbox → File → Import Library   ║
║  2. library_tags.xlsx     →  Open in Numbers / Excel             ║
╚══════════════════════════════════════════════════════════════════╝
""")


if __name__ == "__main__":
    main()
